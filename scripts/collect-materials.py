#!/usr/bin/env python3
"""
collect-materials.py - Phase 1: 装置マニュアル用素材自動収集
Usage: python3 collect-materials.py M1308 "ベーコン原木をハーフカットする装置。伊藤ハム米久プラント柏工場" /tmp/manual-M1308/materials.json
"""

import sys
import os
import json
import glob
import re
import subprocess
from pathlib import Path

DROPBOX = "/Users/daijiromatsuokam1/Machinelab Dropbox/Matsuoka Daijiro"

# ==============================================================================
# 1. プロジェクトフォルダ検索
# ==============================================================================
def find_project_folder(device_number: str) -> str:
    """Dropbox内から装置番号に一致するプロジェクトフォルダを検索"""
    pattern = os.path.join(DROPBOX, f"{device_number}_*")
    folders = glob.glob(pattern)
    if not folders:
        raise FileNotFoundError(f"プロジェクトフォルダが見つかりません: {pattern}")
    if len(folders) > 1:
        print(f"[WARN] 複数フォルダ検出、最初を使用: {folders}", file=sys.stderr)
    return folders[0]


# ==============================================================================
# 2. 部品表Excel → JSON
# ==============================================================================
def parse_bom(project_folder: str) -> list:
    """部品表Excelを読み取りJSON配列に変換"""
    import openpyxl

    bom_patterns = ["部品表/*.xlsx", "BOM/*.xlsx", "*部品表*.xlsx"]
    bom_file = None
    for pat in bom_patterns:
        found = glob.glob(os.path.join(project_folder, pat))
        if found:
            bom_file = found[0]
            break

    if not bom_file:
        print("[WARN] 部品表が見つかりません", file=sys.stderr)
        return []

    print(f"  部品表: {os.path.basename(bom_file)}")
    wb = openpyxl.load_workbook(bom_file, data_only=True)
    ws = wb[wb.sheetnames[0]]  # Sheet1を使用

    parts = []
    for row in ws.iter_rows(min_row=2, values_only=True):  # ヘッダスキップ
        if row[0] is None and row[1] is None:
            continue
        parts.append({
            "number": row[0] if row[0] else "",
            "name": str(row[1]).strip() if row[1] else "",
            "qty": row[2] if row[2] else ""
        })

    print(f"  → {len(parts)}部品抽出")
    return parts


# ==============================================================================
# 3. ニモニック解析
# ==============================================================================
def find_mnemonic(project_folder: str) -> str:
    """ニモニックファイル(.mnm)を再帰検索"""
    found = glob.glob(os.path.join(project_folder, "**/*.mnm"), recursive=True)
    if not found:
        return None
    # Main.mnmを優先
    for f in found:
        if "Main" in os.path.basename(f):
            return f
    return found[0]


def parse_mnemonic(filepath: str) -> dict:
    """ニモニックファイルからI/Oマップ・制御フロー・DM設定を自動抽出"""
    if not filepath:
        print("[WARN] ニモニックファイルが見つかりません", file=sys.stderr)
        return {"io_map": {"inputs": [], "outputs": []}, "dm_settings": [], "safety_interlocks": [], "control_sections": []}

    print(f"  ニモニック: {os.path.basename(filepath)}")

    with open(filepath, "rb") as f:
        raw = f.read()

    # Shift-JIS → UTF-8
    text = raw.decode("shift_jis", errors="replace")
    lines = text.splitlines()

    # デバイスコメントを収集
    device_comments = {}  # addr -> comment
    inputs = []
    outputs = []
    dm_values = {}  # DM addr -> set of compared values
    dm_comments = {}  # DM addr -> comment
    safety_devices = []
    control_sections = []

    # I/Oアドレス範囲の推定（KV-N PLCの一般的な割り当て）
    # R000-R015: 入力リレー, R500-R515: 出力リレー（プロジェクトにより異なる）
    input_range = set()
    output_range = set()

    for line in lines:
        line = line.strip()
        if not line or line.startswith("DEVICE:") or line.startswith(";MODULE"):
            continue

        # セクションヘッダ
        if line.startswith(";<h1/>"):
            section_title = line.replace(";<h1/>", "").strip()
            control_sections.append({"title": section_title, "summary": ""})
            continue

        # コメント抽出
        comment_match = re.search(r";\s*(.+)$", line)
        comment = comment_match.group(1).strip() if comment_match else ""

        # デバイスアドレス抽出
        # パターン: 命令 アドレス [オペランド] ; コメント
        parts = line.split(";")[0].strip().split()
        if len(parts) < 2:
            continue

        instruction = parts[0]
        addr = parts[1]

        # デバイスコメント登録
        if comment and addr not in device_comments:
            device_comments[addr] = comment

        # I/O分類（Rリレーのアドレス範囲で判定）
        r_match = re.match(r"^R(\d+)$", addr)
        if r_match:
            r_num = int(r_match.group(1))
            if r_num < 100:
                input_range.add(addr)
            elif r_num >= 500:
                output_range.add(addr)

        # DM比較値の収集
        dm_match = re.match(r"^DM(\d+)$", addr)
        if dm_match and instruction in ("LD=", "AND=", "OR=", "LD<>", "AND<>", "OR<>", "AND>="):
            dm_addr = addr
            if dm_addr not in dm_values:
                dm_values[dm_addr] = set()
            # 比較値を取得
            if len(parts) >= 3:
                dm_values[dm_addr].add(parts[2])
            if comment:
                dm_comments[dm_addr] = comment

        # OUT命令でDM設定検出
        if instruction == "MOV" and len(parts) >= 3:
            target = parts[2]
            dm_target_match = re.match(r"^DM(\d+)$", target)
            if dm_target_match and comment:
                dm_comments[target] = comment

    # I/Oマップ構築
    for addr in sorted(input_range, key=lambda x: int(re.search(r"\d+", x).group())):
        comment = device_comments.get(addr, "")
        inputs.append({"addr": addr, "name": comment, "type": "input"})

    for addr in sorted(output_range, key=lambda x: int(re.search(r"\d+", x).group())):
        comment = device_comments.get(addr, "")
        outputs.append({"addr": addr, "name": comment, "type": "output"})

    # DM設定の構築
    dm_settings = []
    # DM0-DM2 は制御用（M1308パターン）、それ以上はカウンタ/タイマ値
    for dm_addr in sorted(dm_values.keys()):
        dm_num = int(re.search(r"\d+", dm_addr).group())
        if dm_num > 2:
            continue  # DM3以降はパラメータ系、設定変更対象外
        values = sorted(dm_values[dm_addr])
        comment = dm_comments.get(dm_addr, "")
        dm_settings.append({
            "addr": dm_addr,
            "values": {v: "" for v in values},  # 値の説明はAIが補完
            "desc": comment
        })

    # 安全インターロック抽出
    for addr, comment in device_comments.items():
        if "安全" in comment or "非常" in comment or "カバー" in comment or "ｶﾊﾞｰ" in comment:
            safety_devices.append(f"{comment} ({addr})")

    # 制御セクションにサマリー追加（デバイスコメントから推定）
    for section in control_sections:
        # セクション内の主要デバイスからサマリー推定
        if "投入" in section["title"] and "要求" in section["title"]:
            section["summary"] = "スライサー信号/強制投入の管理"
        elif "運転" in section["title"] and "制御" in section["title"]:
            section["summary"] = "安全条件確認+運転許可"
        elif "FLOW" in section["title"] or "フロー" in section["title"]:
            section["summary"] = "搬送→カット→投入完了シーケンス"
        elif "出力" in section["title"]:
            section["summary"] = "モーター駆動+ランプ+信号出力"

    result = {
        "io_map": {"inputs": inputs, "outputs": outputs},
        "dm_settings": dm_settings,
        "safety_interlocks": safety_devices,
        "control_sections": control_sections,
        "device_comments": device_comments,
        "raw_line_count": len(lines)
    }
    print(f"  → 入力{len(inputs)}点, 出力{len(outputs)}点, DM設定{len(dm_settings)}項, 安全装置{len(safety_devices)}件")
    return result


# ==============================================================================
# 4. 信号仕様PDF → テキスト
# ==============================================================================
def extract_signal_spec(project_folder: str) -> str:
    """信号仕様PDFをテキスト抽出"""
    patterns = [
        "**/*信号*.*pdf", "**/*信号やり取り*.*pdf", "**/*signal*.*pdf"
    ]
    for pat in patterns:
        found = glob.glob(os.path.join(project_folder, pat), recursive=True)
        if found:
            pdf_path = found[0]
            print(f"  信号仕様: {os.path.basename(pdf_path)}")
            try:
                result = subprocess.run(
                    ["pdftotext", "-layout", pdf_path, "-"],
                    capture_output=True, text=True, timeout=30
                )
                if result.returncode == 0 and result.stdout.strip():
                    print(f"  → {len(result.stdout)}文字抽出")
                    return result.stdout
            except Exception as e:
                print(f"  [WARN] pdftotext失敗: {e}", file=sys.stderr)
    print("  [WARN] 信号仕様PDFが見つかりません", file=sys.stderr)
    return ""


# ==============================================================================
# 5. 電装図PDF → テキスト
# ==============================================================================
def extract_electrical(project_folder: str, device_number: str) -> str:
    """電装図PDFをテキスト抽出"""
    # 内海制御フォルダ内のPDF図面を優先検索
    patterns = [
        f"**/001_PDF*/{device_number}-E-*.pdf",
        f"**/{device_number}-E-003_*.pdf",
        f"**/{device_number}-E-S*.pdf",
        f"**/{device_number}-E-P*.pdf",
        "**/電装図*.pdf"
    ]
    found_set = set()
    pdf_files = []
    for pat in patterns:
        found = glob.glob(os.path.join(project_folder, pat), recursive=True)
        for f in found:
            if f not in found_set:
                found_set.add(f)
                pdf_files.append(f)

    # マニュアルに有用な図面を優先: PLC入出力回路、外部配線、PLC構成図
    priority_keywords = ["S061", "S081", "S015", "S050", "S001", "E-003"]
    priority_files = []
    other_files = []
    for f in pdf_files:
        basename = os.path.basename(f)
        if any(kw in basename for kw in priority_keywords):
            priority_files.append(f)
        else:
            other_files.append(f)

    # 優先ファイルを先に、残りは後ろに
    ordered_files = priority_files + other_files

    texts = []
    for pdf_path in ordered_files:
        try:
            result = subprocess.run(
                ["pdftotext", "-layout", pdf_path, "-"],
                capture_output=True, text=True, timeout=30
            )
            if result.returncode == 0 and result.stdout.strip():
                texts.append(f"=== {os.path.basename(pdf_path)} ===\n{result.stdout}")
        except Exception:
            pass
    if texts:
        combined = "\n".join(texts)
        print(f"  電装図: {len(texts)}ファイル, {len(combined)}文字")
        return combined
    print("  [WARN] 電装図PDFが見つかりません", file=sys.stderr)
    return ""


# ==============================================================================
# 6. 画像パス収集
# ==============================================================================
def collect_images(project_folder: str, device_number: str) -> dict:
    """組立図・装置画像のパスを収集"""
    result = {"assembly_drawing": "", "device_images": []}

    # 組立図 (JPG or PDF)
    asm_patterns = [
        f"{device_number}-01-0000_*.JPG",
        f"{device_number}-01-0000_*.jpg",
        f"{device_number}-01-0000_*.pdf",
    ]
    for pat in asm_patterns:
        found = glob.glob(os.path.join(project_folder, pat))
        if found:
            result["assembly_drawing"] = found[0]
            print(f"  組立図: {os.path.basename(found[0])}")
            break

    # 装置画像
    img_dirs = ["装置画像", "イメージ"]
    for d in img_dirs:
        img_dir = os.path.join(project_folder, d)
        if os.path.isdir(img_dir):
            imgs = sorted(glob.glob(os.path.join(img_dir, "*.jpg")) +
                          glob.glob(os.path.join(img_dir, "*.JPG")) +
                          glob.glob(os.path.join(img_dir, "*.png")))
            result["device_images"].extend(imgs)

    print(f"  装置画像: {len(result['device_images'])}枚")
    return result


# ==============================================================================
# 7. 仕様情報の推定
# ==============================================================================
def extract_specs(parts_list: list, plc_data: dict) -> dict:
    """部品表とPLCデータから主要仕様を推定"""
    motors = []
    sensors = []

    for part in parts_list:
        name = str(part.get("name", ""))
        qty = part.get("qty", 1)

        # モーター検出
        if "モーター" in name or "ギヤードモーター" in name or "FPW" in name:
            model_match = re.search(r"(FPW\S+|BLV\S+|BXM\S+)", name)
            model = model_match.group(1) if model_match else name
            motors.append({"name": name, "model": model, "maker": "オリエンタルモーター", "qty": qty})

        # センサー検出
        if "センサー" in name or "センサ" in name or "E2E" in name or "LR-" in name:
            model_match = re.search(r"(E2E\S+|LR-\S+|FU-\S+)", name)
            model = model_match.group(1) if model_match else name
            maker = "オムロン" if "E2E" in name else "キーエンス" if "LR-" in name else ""
            sensors.append({"name": name, "model": model, "maker": maker, "qty": qty})

    # PLCのI/Oからもセンサー情報補完
    if plc_data and plc_data.get("io_map"):
        for inp in plc_data["io_map"].get("inputs", []):
            inp_name = inp.get("name", "")
            if ("PX" in inp_name or "センサ" in inp_name or "PH" in inp_name) and \
               not any(inp_name in s.get("name", "") for s in sensors):
                pass  # 部品表で既にカバーされているはず

    return {
        "motors": motors,
        "sensors": sensors
    }


# ==============================================================================
# 8. 説明文パース
# ==============================================================================
def parse_description(desc: str) -> tuple:
    """説明文から装置説明と納品先を分離"""
    # "ベーコン原木をハーフカットする装置。伊藤ハム米久プラント柏工場"
    # → ("ベーコン原木をハーフカットする装置", "伊藤ハム米久プラント柏工場")
    parts = re.split(r"[。．.、]", desc, maxsplit=1)
    if len(parts) == 2:
        return parts[0].strip(), parts[1].strip()
    return desc.strip(), ""


# ==============================================================================
# Main
# ==============================================================================
def main():
    if len(sys.argv) < 4:
        print(f"Usage: {sys.argv[0]} <device_number> <description> <output_json>")
        sys.exit(1)

    device_number = sys.argv[1]
    description = sys.argv[2]
    output_path = sys.argv[3]

    device_desc, customer_raw = parse_description(description)

    print(f"=== 素材収集: {device_number} ===")

    # 1. プロジェクトフォルダ
    project_folder = find_project_folder(device_number)
    print(f"  フォルダ: {os.path.basename(project_folder)}")

    # フォルダ名から装置名・納品先を推定
    folder_name = os.path.basename(project_folder)
    # "M1308_伊藤ハム米久プラント柏工場_ベーコンハーフカット装置"
    folder_parts = folder_name.split("_", 1)
    if len(folder_parts) > 1:
        remaining = folder_parts[1]
        # 最後のパートが装置名
        remaining_parts = remaining.rsplit("_", 1)
        if len(remaining_parts) == 2:
            customer_from_folder = remaining_parts[0]
            device_name_from_folder = remaining_parts[1]
        else:
            customer_from_folder = ""
            device_name_from_folder = remaining
    else:
        customer_from_folder = ""
        device_name_from_folder = ""

    # 納品先: 引数 > フォルダ名
    customer = customer_raw if customer_raw else customer_from_folder

    # 2. 部品表
    print("[1/6] 部品表...")
    parts_list = parse_bom(project_folder)

    # 3. ニモニック
    print("[2/6] ニモニック...")
    mnm_path = find_mnemonic(project_folder)
    plc_data = parse_mnemonic(mnm_path)

    # 4. 信号仕様
    print("[3/6] 信号仕様...")
    signal_text = extract_signal_spec(project_folder)

    # 5. 電装図
    print("[4/6] 電装図...")
    electrical_text = extract_electrical(project_folder, device_number)

    # 6. 画像
    print("[5/6] 画像...")
    images = collect_images(project_folder, device_number)

    # 7. 仕様推定
    print("[6/6] 仕様推定...")
    specs = extract_specs(parts_list, plc_data)

    # 出力JSON
    materials = {
        "device_number": device_number,
        "device_name": device_name_from_folder or "装置",
        "customer": customer,
        "customer_formal": f"{customer} 様" if customer else "",
        "manufacturer": "株式会社キカイラボ",
        "manufacturer_tel": "047-404-9713",
        "manufacturer_email": "info@machinelab.co.jp",
        "description": device_desc,
        "date": __import__("datetime").datetime.now().strftime("%Y年%m月"),
        "project_folder": project_folder,
        "parts_list": parts_list,
        "plc": plc_data,
        "signal_spec_text": signal_text,
        "electrical_text": electrical_text[:5000] if electrical_text else "",  # 長すぎる場合は切り詰め
        "assembly_drawing": images["assembly_drawing"],
        "device_images": images["device_images"],
        "specs": specs
    }

    # 出力ディレクトリ作成
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(materials, f, ensure_ascii=False, indent=2)

    print(f"\n✅ 素材収集完了: {output_path}")
    print(f"   部品: {len(parts_list)}, PLC I/O: {len(plc_data['io_map']['inputs'])}入力/{len(plc_data['io_map']['outputs'])}出力")
    print(f"   信号仕様: {'あり' if signal_text else 'なし'}, 電装図: {'あり' if electrical_text else 'なし'}")
    print(f"   画像: {len(images['device_images'])}枚, 組立図: {'あり' if images['assembly_drawing'] else 'なし'}")


if __name__ == "__main__":
    main()
